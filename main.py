import requests
import json
import yfinance as yf
import pandas as pd
import re
import openpyxl
from datetime import datetime
import os
import argparse

CONFIG_PATH = 'config/config.json'
DATA_PATH = 'data'

sheets = [
    'BALANCE_SHEET',
    'INCOME_STATEMENT',
    'CASH_FLOW'
]

def get_api_key(config_path):
    """Retrieves the API key from a configuration file."""
    try:
        with open(config_path, 'r') as file:
            config_data = json.load(file)
            api_key = config_data.get('api_key')
            if not api_key:
                raise ValueError("API key not found in configuration file.")
            return api_key
    except (FileNotFoundError, json.JSONDecodeError) as e:
        raise RuntimeError(f"Error reading API key: {e}")

def get_statements(symbol, no_fetch):
    """Fetches financial statements for a given stock symbol and saves them as JSON files."""
    
    if no_fetch == True:
        count = 0
        for i in range(len(sheets)):
            if os.path.exists(f'{DATA_PATH}/{symbol}_{sheets[i]}.json'):
                count += 1

        if count == len(sheets):
            print("Skipped fetching statements, files found, exporting...")
            return 0
        else:
            raise FileNotFoundError(f"Could not find JSON files.")
    
    try:
        api_key = get_api_key(CONFIG_PATH)
    except RuntimeError as e:
        print(e)
        return

    if api_key == 'INSERT YOUR KEY HERE':
        print("Please provide an API key. (config/config.json)")
        return

    for sheet in sheets:
        url = f'https://www.alphavantage.co/query?function={sheet}&symbol={symbol}&apikey={api_key}'
        response = requests.get(url)
        data = response.json()

        # Check for API rate limit message
        if "Information" in data:
            print(f"API Limit Reached: {data['Information']}")
            return -1 # Stop execution if rate limit is reached
        
        # Ensure output directory exists
        os.makedirs(DATA_PATH, exist_ok=True)

        # Save the response data to a JSON file
        file_path = os.path.join(DATA_PATH, f"{symbol}_{sheet}.json")
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)

        print(f"Saved: {file_path}")  # Logging success
        
    return 0


def get_info(ticker_symbol):
    try:
        ticker = yf.Ticker(ticker_symbol)
        company_name = ticker.info.get("longName", "Unknown Company Name")
        sector = ticker.info.get("sector", "Unknown Sector")
        industry = ticker.info.get("industry", "Unknown Industry")
        website = ticker.info.get("website", "N/A")

        return {
            "company_name": company_name,
            "sector": sector,
            "industry": industry,
            "website": website
        }
        
    except:
        print("Unable to fetch company data.")
        return {
            "company_name": "Unknown Company Name",
            "sector": "Unknown Sector",
            "industry": "Unknown Industry",
            "website": "N/A"
            }


def process_statement(data_path, years):
    with open(data_path, 'r') as file:
        data = json.load(file)
    quarterly = data["quarterlyReports"]


    periods = []
    num_of_periods = years * 4
    for i in range(num_of_periods):
        periods.append(quarterly[i]["fiscalDateEnding"])

    if quarterly:
        keys_list = list(quarterly[0].keys())
    
    title_case_keys = []
    for i in range(len(keys_list)):
        current = re.sub(r'([a-z])([A-Z])', r'\1 \2', keys_list[i])
        title_case_keys.append(current.title()) 

    statement = {}
    statement["Fiscal Date Ending"] = title_case_keys[2:] # exclude fiscalDateEnding and reportedCurrency to match the length of the other lists below
    for period in range(len(periods)):
        values = []
        for key in range(2, len(keys_list)): # from 2 to only get the numbers
            if quarterly[period][keys_list[key]] != 'None':
                values.append(quarterly[period][keys_list[key]])
            else:
                values.append(0)
        statement[quarterly[period][keys_list[0]]] = values

    df = pd.DataFrame(statement)
    df = df.set_index("Fiscal Date Ending")
    df = df.iloc[:, ::-1]
    return df


def export(data_path, symbol, years):
    # Retrieve company information
    company_info = get_info(symbol)

    # Create the cover sheet as a DataFrame
    cover_data = {
        "Company Name": [company_info['company_name']],
        "Sector": [company_info['sector']],
        "Industry": [company_info['industry']],
        "Website": [company_info['website']],
        "Retrieved": [datetime.now()]
    }
    cover_df = pd.DataFrame(cover_data)

    # Process other statements
    balance_sheet = process_statement(f'{data_path}/{symbol}_{sheets[0]}.json', years)
    income_statement = process_statement(f'{data_path}/{symbol}_{sheets[1]}.json', years)
    cash_flow = process_statement(f'{data_path}/{symbol}_{sheets[2]}.json', years)

    os.makedirs('output', exist_ok=True)
    with pd.ExcelWriter(f'output/{symbol}_statements.xlsx', engine='openpyxl') as writer:
        
        # Write the cover sheet first
        cover_df.to_excel(writer, sheet_name="Cover Sheet", index=False)
        
        balance_sheet = balance_sheet.apply(pd.to_numeric)
        income_statement = income_statement.apply(pd.to_numeric)
        cash_flow = cash_flow.apply(pd.to_numeric)

        balance_sheet.to_excel(writer, sheet_name='Balance Sheet')
        income_statement.to_excel(writer, sheet_name='Income Statement')
        cash_flow.to_excel(writer, sheet_name='Cash Flow')
        
        workbook = writer.book
        
        custom_format = '0,,;(0,,)'
        short_date_format = 'MM/DD/YYYY'

        # Loop through each sheet and apply formatting
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Apply custom format to each cell in the sheet (excluding headers)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.number_format = custom_format
            
            if sheet_name == "Cover Sheet":
                retrieved_cell = sheet.cell(row=2, column=5)
                retrieved_cell.number_format = short_date_format  # Apply the date format

                for col_num, column in enumerate(sheet.columns, 1):
                    column_letter = openpyxl.utils.get_column_letter(col_num)  # Convert column number to letter
                    sheet.column_dimensions[column_letter].width = 25
            
            # Set column width for all columns except the first one (14 pixels)
            if sheet_name != "Cover Sheet":
                for col_num, column in enumerate(sheet.columns, 1):
                    if col_num != 1:  # Skip the first column
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        sheet.column_dimensions[column_letter].width = 13
                    else:
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        sheet.column_dimensions[column_letter].width = 35


def validate_ticker(symbol):
    """Validates and checks if the ticker exists using Yahoo Finance."""
    symbol = symbol.strip().upper()
    if not re.match(r"^[A-Z0-9]{1,5}$", symbol):
        raise argparse.ArgumentTypeError(f"Invalid ticker: '{symbol}'. Must be 1-5 letters/numbers.")

    # Check if the ticker exists
    if not yf.Ticker(symbol).history(period="1d").empty:
        return symbol
    else:
        raise argparse.ArgumentTypeError(f"Ticker '{symbol}' not found in Yahoo Finance.")


def validate_years(value):
    try:
        value = int(value)  # Convert input to integer
        if 1 <= value <= 15:
            return value
        else:
            raise argparse.ArgumentTypeError("Invalid value: Must be an integer between 1 and 15.")
    except ValueError:
        raise argparse.ArgumentTypeError("Invalid value: Must be an integer (1 - 15).")



def main(data_path, symbol, years, no_fetch):
    print(f"Retrieving statements for {symbol} for the last {years} year(s).")
    gs = get_statements(symbol, no_fetch)
    if gs != 0:
        return

    export(data_path, symbol, years)
    print(f"Excel file saved! Location: output/{symbol}_statements.xlsx")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export Financial Statements for a given Company")
    
    parser.add_argument(
        'symbol',
        type=validate_ticker,
        help="The company's ticker (e.g. AAPL, MSFT)"
        )
    
    parser.add_argument(
        'years',
        nargs='?',
        default=5,
        type=validate_years,
        help="Optional argument, the number of years to export statements for (1 - 15). Default value is 5."
    )

    parser.add_argument(
        '--no-fetch',
        action='store_true',
        help=
            """Skip fetching the data from the API to save on the number of calls you have to make.
            Useful when you have already exported into excel and you want to change the number of years the sheet contains.
            """
    )

    args = parser.parse_args()    
    main(DATA_PATH, args.symbol, args.years, args.no_fetch)